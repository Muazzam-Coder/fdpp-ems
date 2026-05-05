from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters import rest_framework as filters
from django.contrib.auth.models import User
from .models import Employee, Attendance, PaidLeave, Shift, UserAccessLevel
from .serializers import (
    EmployeeSerializer, AttendanceSerializer, PaidLeaveSerializer, 
    ShiftSerializer, UserSerializer, UserAccessLevelSerializer,
    CreateAdminManagerSerializer, RegisterSerializer
)
from django.db.models import Sum, Count, Q, Avg
from datetime import datetime, timedelta, date, time
from django.utils import timezone
from fdpp_ems import settings
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from io import BytesIO
from django.http import HttpResponse
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

# Permission check: Only admins can create admin/manager
def is_admin(user):
    """Check if user has admin role"""
    try:
        return user.access_level.role == 'admin'
    except (AttributeError, UserAccessLevel.DoesNotExist):
        return False


def format_hours_display(hours_value):
    if hours_value is None:
        return '0h 0m'

    total_minutes = max(0, int(round(float(hours_value) * 60)))
    hours, minutes = divmod(total_minutes, 60)
    return f'{hours}h {minutes}m'


def build_absent_entries(report_date, request=None):
    """Build synthetic absent rows for active employees who missed the shift."""
    current_time = datetime.now().time()
    today = timezone.now().date()

    active_employees = Employee.objects.filter(status='active')
    attendances = Attendance.objects.filter(date=report_date)
    present_employee_ids = set(attendances.values_list('employee', flat=True).distinct())

    absent_entries = []
    pending_count = 0

    for employee in active_employees:
        if employee.emp_id in present_employee_ids:
            continue

        # If employee has no defined shift times, decide pending vs absent based on date
        if not employee.start_time or not employee.end_time:
            if report_date < today:
                on_leave = PaidLeave.objects.filter(
                    employee=employee,
                    approved=True,
                    start_time__date__lte=report_date,
                    end_time__date__gte=report_date,
                ).exists()
                status_val = "on_leave" if on_leave else "absent"
                absent_entries.append({
                    "id": None,
                    "employee": employee.emp_id,
                    "employee_name": employee.name,
                    "date": report_date,
                    "check_in": None,
                    "check_out": None,
                    "message_late": None,
                    "status": status_val,
                    "total_hours": "0h 0m",
                    "is_late": False,
                    "created_at": None,
                    "updated_at": None,
                })
            else:
                pending_count += 1
            continue

        if report_date < today or (report_date == today and current_time >= employee.end_time):
            # Check if employee has an approved paid leave covering this date
            on_leave = PaidLeave.objects.filter(
                employee=employee,
                approved=True,
                start_time__date__lte=report_date,
                end_time__date__gte=report_date,
            ).exists()

            status_val = "on_leave" if on_leave else "absent"
            absent_entries.append({
                "id": None,
                "employee": employee.emp_id,
                "employee_name": employee.name,
                "date": report_date,
                "check_in": None,
                "check_out": None,
                "message_late": None,
                "status": status_val,
                "total_hours": "0h 0m",
                "is_late": False,
                "created_at": None,
                "updated_at": None,
            })
        else:
            pending_count += 1

    return absent_entries, pending_count

class IsAdmin(IsAuthenticated):
    """Permission class for admin access"""
    def has_permission(self, request, view):
        return super().has_permission(request, view) and is_admin(request.user)

# Authentication ViewSet
class AuthViewSet(viewsets.ViewSet):
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['post'])
    def register(self, request):
        """Register a new user and create employee profile with image upload"""
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            result = serializer.save()
            user = result['user']
            employee = result['employee']
            
            return Response({
                "message": "User registered successfully",
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "first_name": user.first_name,
                    "last_name": user.last_name
                },
                "employee": {
                    "emp_id": employee.emp_id,
                    "name": employee.name,
                    "designation": employee.designation,
                    "phone": employee.phone,
                    "CNIC": employee.CNIC,
                    "shift_type": employee.shift_type,
                    "start_time": employee.start_time.strftime('%H:%M:%S') if employee.start_time else None,
                    "end_time": employee.end_time.strftime('%H:%M:%S') if employee.end_time else None,
                    "profile_img": f"http://{settings.SERVER_IP}:{settings.SERVER_PORT}{employee.profile_img.url}" if employee.profile_img else None
                }
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def create_admin_manager(self, request):
        """Create a new admin or manager user with optional profile image (admin only)"""
        serializer = CreateAdminManagerSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            
            response_data = {
                "message": f"User created successfully as {serializer.validated_data['role']}",
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "role": serializer.validated_data['role'],
                    "profile_img": None
                }
            }
            
            # Add profile image URL if it exists
            try:
                user_profile = user.profile
                if user_profile.profile_img:
                    response_data["user"]["profile_img"] = request.build_absolute_uri(user_profile.profile_img.url)
            except:
                pass
            
            return Response(response_data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def login(self, request):
        """User login endpoint"""
        from django.contrib.auth import authenticate
        username = request.data.get('username')
        password = request.data.get('password')
        
        user = authenticate(username=username, password=password)
        if user:
            try:
                # Fetch access level directly from database to get latest data
                access_level = UserAccessLevel.objects.get(user=user)
                return Response({
                    "message": "Login successful",
                    "user_id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "role": access_level.role
                }, status=status.HTTP_200_OK)
            except UserAccessLevel.DoesNotExist:
                # Try to get employee profile if user has one
                try:
                    employee = user.employee_profile
                    return Response({
                        "message": "Login successful",
                        "user_id": user.id,
                        "username": user.username,
                        "email": user.email,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "emp_id": employee.emp_id,
                        "name": employee.name,
                        "role": "employee"
                    }, status=status.HTTP_200_OK)
                except Employee.DoesNotExist:
                    return Response(
                        {"error": "User profile not found"},
                        status=status.HTTP_404_NOT_FOUND
                    )
        return Response(
            {"error": "Invalid credentials"},
            status=status.HTTP_401_UNAUTHORIZED
        )


class UserAccessLevelViewSet(viewsets.ModelViewSet):
    """Manage user access levels (admin/manager)"""
    queryset = UserAccessLevel.objects.all()
    serializer_class = UserAccessLevelSerializer
    permission_classes = [IsAdmin]
    
    @action(detail=False, methods=['get'])
    def admins(self, request):
        """Get all admin users"""
        admins = UserAccessLevel.objects.filter(role='admin')
        serializer = self.get_serializer(admins, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def managers(self, request):
        """Get all manager users"""
        managers = UserAccessLevel.objects.filter(role='manager')
        serializer = self.get_serializer(managers, many=True)
        return Response(serializer.data)


# Filtering logic for Attendance
class AttendanceFilter(filters.FilterSet):
    date_from = filters.DateFilter(field_name="date", lookup_expr='gte')
    date_to = filters.DateFilter(field_name="date", lookup_expr='lte')
    employee = filters.CharFilter(field_name="employee__emp_id")
    status = filters.CharFilter(field_name="status")

    class Meta:
        model = Attendance
        fields = ['employee', 'date_from', 'date_to', 'status']

class EmployeeFilter(filters.FilterSet):
    employee_id = filters.NumberFilter(field_name="emp_id", lookup_expr='exact')
    employee_name = filters.CharFilter(field_name="name", lookup_expr='icontains')
    status = filters.CharFilter(field_name="status")

    class Meta:
        model = Employee
        fields = ['employee_id', 'employee_name', 'status', 'shift_type']

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    filterset_class = EmployeeFilter
    filter_backends = (filters.DjangoFilterBackend,)
    lookup_field = 'emp_id'  # Use emp_id for URL lookups like /employees/EMP001/
    permission_classes = [IsAuthenticated]

    def update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return super().update(request, *args, **kwargs)

    @action(detail=True, methods=['get'])
    def calculate_payout(self, request, emp_id=None):
        """Calculates salary based on attendance and hourly rate."""
        employee = self.get_object()
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if not start_date or not end_date:
            return Response(
                {"error": "Please provide start_date and end_date (YYYY-MM-DD)"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        attendances = Attendance.objects.filter(
            employee=employee, 
            date__range=[start_date, end_date]
        )
        
        total_worked_hours = round(sum(att.total_hours for att in attendances), 2)
        payout = float(total_worked_hours) * float(employee.hourly_rate or 0)

        return Response({
            "employee_id": employee.emp_id,
            "employee_name": employee.name,
            "period": f"{start_date} to {end_date}",
            "total_hours": format_hours_display(total_worked_hours),
            "total_hours_value": total_worked_hours,
            "hourly_rate": float(employee.hourly_rate or 0),
            "total_payout": float(payout)
        })

    @action(detail=True, methods=['get'])
    def attendance_report(self, request, emp_id=None):
        """Get attendance report for an employee with filters"""
        employee = self.get_object()
        period = request.query_params.get('period', 'month')  # day, week, month, custom
        
        today = timezone.now().date()
        
        if period == 'day':
            start_date = today
            end_date = today
        elif period == 'week':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == 'month':
            start_date = date(today.year, today.month, 1)
            if today.month == 12:
                end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
        elif period == 'custom':
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            if not start_date_str or not end_date_str:
                return Response(
                    {"error": "Please provide start_date and end_date for custom period"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            return Response(
                {"error": "Invalid period. Use: day, week, month, or custom"},
                status=status.HTTP_400_BAD_REQUEST
            )

        attendances = Attendance.objects.filter(
            employee=employee,
            date__range=[start_date, end_date]
        ).order_by('-date')

        total_hours = round(sum(att.total_hours for att in attendances), 2)
        total_days = attendances.count()
        late_count = attendances.filter(status='late').count()
        on_time_count = attendances.filter(status='on_time').count()

        return Response({
            "employee_id": employee.emp_id,
            "employee_name": employee.name,
            "period": f"{start_date} to {end_date}",
            "total_days_worked": total_days,
            "total_hours": format_hours_display(total_hours),
            "total_hours_value": total_hours,
            "on_time": on_time_count,
            "late": late_count,
            "attendance_records": AttendanceSerializer(attendances, many=True).data
        })

    @action(detail=False, methods=['get'])
    def active_employees(self, request):
        """Get list of active employees"""
        employees = Employee.objects.filter(status='active')
        serializer = self.get_serializer(employees, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def employee_stats(self, request):
        """Get overall employee statistics"""
        total_employees = Employee.objects.count()
        active_employees = Employee.objects.filter(status='active').count()
        inactive_employees = Employee.objects.filter(status='inactive').count()
        
        today = timezone.now().date()
        present_today = Attendance.objects.filter(date=today).values('employee').distinct().count()
        
        return Response({
            "total_employees": total_employees,
            "active_employees": active_employees,
            "inactive_employees": inactive_employees,
            "present_today": present_today
        })

    @action(detail=False, methods=['get', 'post'], url_path='relatives')
    def relatives(self, request):
        """Get or set relatives for an employee.

        GET params:
        - emp_id (required): employee emp_id to query
        - transitive (optional): 'true' to return full connected relatives graph
        - name (optional): filter relatives by name (icontains)

        POST body (json):
        - emp_id (required): employee emp_id to update
        - relatives: comma-separated string of emp_id values OR list of ints

        Response: single variable `relatives` containing list of relatives (dicts).
        """
        if request.method == 'GET':
            emp_id = request.query_params.get('emp_id') or request.query_params.get('employee')
        else:
            emp_id = request.data.get('emp_id')

        if not emp_id:
            return Response({"error": "Please provide emp_id"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(emp_id=emp_id)
        except Employee.DoesNotExist:
            return Response({"error": f"Employee with id {emp_id} not found"}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'POST':
            # accept comma-separated string or list of emp_id values or numeric PKs
            raw = request.data.get('relatives', '')
            rel_inputs = []
            if isinstance(raw, str):
                raw = raw.strip()
                if raw:
                    rel_inputs = [x.strip() for x in raw.split(',') if x.strip()]
            elif isinstance(raw, (list, tuple)):
                rel_inputs = [x for x in raw if x is not None]

            # Normalize inputs to emp_id values where possible
            resolved_emp_ids = []
            for v in rel_inputs:
                vs = str(v).strip()
                # try emp_id lookup first
                try:
                    obj = Employee.objects.get(emp_id=vs)
                    resolved_emp_ids.append(obj.emp_id)
                    continue
                except Employee.DoesNotExist:
                    pass
                # try pk lookup
                if vs.isdigit():
                    try:
                        obj = Employee.objects.get(pk=int(vs))
                        resolved_emp_ids.append(obj.emp_id)
                        continue
                    except Employee.DoesNotExist:
                        pass

            # Save comma-separated into legacy `relative` field (store emp_id values)
            employee.relative = ','.join(resolved_emp_ids)
            employee.save()

            # Update M2M relations using emp_id values
            relatives_qs = Employee.objects.filter(emp_id__in=resolved_emp_ids)
            employee.relatives.set(relatives_qs)

            # Prepare response (use emp_id as the identifier)
            data_qs = relatives_qs.order_by('emp_id')
            data = [{"emp_id": r.emp_id, "name": r.name} for r in data_qs]
            return Response({"relatives": data})

        # GET handling
        transitive = str(request.query_params.get('transitive', '')).lower() in ('1', 'true', 'yes')
        name_filter = request.query_params.get('name')

        if not transitive:
            relatives_qs = employee.relatives.all()
        else:
            visited = set()
            queue = [employee]
            visited.add(employee.pk)
            collected = set()
            while queue:
                current = queue.pop(0)
                for r in current.relatives.all():
                    if r.pk not in visited:
                        visited.add(r.pk)
                        queue.append(r)
                    if r.pk != employee.pk:
                        collected.add(r.pk)
            relatives_qs = Employee.objects.filter(pk__in=collected)

        if name_filter:
            relatives_qs = relatives_qs.filter(name__icontains=name_filter)

        data = [{"emp_id": r.emp_id, "name": r.name} for r in relatives_qs.order_by('emp_id')]
        return Response({"relatives": data})

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all().order_by('-date')
    serializer_class = AttendanceSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = AttendanceFilter
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        """Return attendance list with filters (date, date_from/date_to, employee, status).

        - If `date` is provided (single day) we include synthetic absent rows after shift end.
        - Otherwise use DRF filters (`AttendanceFilter`) and support pagination.
        """
        date_str = request.query_params.get('date')

        # Start with DRF-filtered queryset so filters like employee, date_from/date_to, status apply
        qs = self.filter_queryset(self.get_queryset())

        # Single-date behavior: include absent entries and day-level summary
        if date_str:
            try:
                report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({"error": "Invalid date format"}, status=status.HTTP_400_BAD_REQUEST)

            attendances = qs.filter(date=report_date).order_by('employee', 'check_in')
            serializer = AttendanceSerializer(attendances, many=True, context={'request': request})
            absent_entries, pending_count = build_absent_entries(report_date, request=request)

            results = list(serializer.data) + absent_entries
            present_count = attendances.values('employee').distinct().count()
            absent_count = len(absent_entries)
            late_count = attendances.filter(status='late').values('employee').distinct().count()
            on_time_count = max(0, present_count - late_count)
            total_hours = round(sum(att.total_hours for att in attendances), 2)

            return Response({
                "date": report_date,
                "present": present_count,
                "absent": absent_count,
                "pending": pending_count,
                "on_time": on_time_count,
                "late": late_count,
                "total_hours": format_hours_display(total_hours),
                "total_hours_value": total_hours,
                "count": len(results),
                "results": results,
            })

        # Multi-day / filtered list: use filtered queryset and support pagination
        queryset = qs.order_by('-date', 'employee', 'check_in')

        # Compute aggregates on full queryset (not just page)
        total_hours = round(sum(att.total_hours for att in queryset), 2)
        present_count = queryset.values('employee').distinct().count()
        total_count = queryset.count()

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = AttendanceSerializer(page, many=True, context={'request': request})
            paginated = self.get_paginated_response(serializer.data).data
            # augment paginated response with summary fields
            paginated.update({
                "present": present_count,
                "total_hours": format_hours_display(total_hours),
                "total_hours_value": total_hours,
                "total_count": total_count,
            })
            return Response(paginated)

        serializer = AttendanceSerializer(queryset, many=True, context={'request': request})
        return Response({
            "present": present_count,
            "total_hours": format_hours_display(total_hours),
            "total_hours_value": total_hours,
            "count": total_count,
            "results": serializer.data,
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export attendance matrix to Excel: rows=dates, columns=employees.

        Query params:
        - date_from (YYYY-MM-DD) required
        - date_to (YYYY-MM-DD) required
        - employees: optional comma-separated emp_id list; defaults to active employees
        - employee: optional single emp_id to export for a specific employee
        """
        if openpyxl is None:
            return Response({"error": "openpyxl is required to export Excel. Please pip install openpyxl."}, status=500)

        start = request.query_params.get('date_from')
        end = request.query_params.get('date_to')
        if not start or not end:
            return Response({"error": "Please provide date_from and date_to in YYYY-MM-DD"}, status=400)
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date()
            end_date = datetime.strptime(end, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        # Accept either `employee` (single emp_id) or `employees` (comma-separated)
        emp_single = request.query_params.get('employee')
        emp_param = request.query_params.get('employees')
        if emp_single and str(emp_single).strip().isdigit():
            employees = list(Employee.objects.filter(emp_id=int(emp_single)).order_by('emp_id'))
        elif emp_param:
            emp_ids = [int(x.strip()) for x in emp_param.split(',') if x.strip().isdigit()]
            employees = list(Employee.objects.filter(emp_id__in=emp_ids).order_by('emp_id'))
        else:
            employees = list(Employee.objects.filter(status='active').order_by('emp_id'))

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance'

        # Header row: Date + for each employee a single column with "CheckIn - CheckOut"
        headers = ['Date']
        for e in employees:
            label = e.name or str(e.emp_id)
            headers.append(f"{label} (In - Out)")

        for col_idx, head in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=head)

        # Fill rows per date
        row = 2
        delta = (end_date - start_date).days
        # Use current time to decide whether a day should be considered absent or still pending
        current_time = datetime.now().time()
        today = timezone.now().date()

        for d_offset in range(delta + 1):
            rdate = start_date + timedelta(days=d_offset)
            ws.cell(row=row, column=1, value=rdate.strftime('%Y-%m-%d'))

            # for each employee write a single cell with "CheckIn - CheckOut"
            for idx, emp in enumerate(employees):
                col = 2 + idx
                # Find attendances for this employee on the row date, either by stored
                # `date` or by the `check_in` datetime falling on that date.
                atts = Attendance.objects.filter(employee=emp).filter(
                    Q(date=rdate) | Q(check_in__date=rdate)
                )

                # Prefer attendances whose actual `check_in` falls on this date to avoid
                # picking an unrelated open record from another day.
                day_atts = atts.filter(check_in__date=rdate).order_by('check_in')
                if day_atts.exists():
                    atts = day_atts
                else:
                    atts = atts.order_by('check_in')

                check_in_val = "--:--"
                check_out_val = "--:--"
                if atts.exists():
                    first = atts.first()
                    last = atts.last()
                    if first.check_in:
                        try:
                            check_in_val = first.check_in.strftime('%I:%M %p')
                        except Exception:
                            check_in_val = str(first.check_in)
                    if last.check_out:
                        try:
                            check_out_val = last.check_out.strftime('%I:%M %p')
                        except Exception:
                            check_out_val = str(last.check_out)

                if atts.exists():
                    combined = f"{check_in_val} - {check_out_val}"
                else:
                    # If no attendance, check for approved paid leave covering the date
                    leave = PaidLeave.objects.filter(
                        employee=emp,
                        approved=True,
                        start_time__date__lte=rdate,
                        end_time__date__gte=rdate,
                    ).first()
                    if leave:
                        combined = "Leave"
                    else:
                        # If employee has no defined shift times, treat past dates as Absent
                        if not emp.start_time or not emp.end_time:
                            if rdate < today:
                                combined = "Absent"
                            else:
                                combined = f"{check_in_val} - {check_out_val}"
                        else:
                            # If the date is past (or today past shift end) mark Absent
                            if rdate < today or (rdate == today and current_time >= emp.end_time):
                                combined = "Absent"
                            else:
                                # Future or pending day: keep placeholder
                                combined = f"{check_in_val} - {check_out_val}"

                ws.cell(row=row, column=col, value=combined)

            row += 1

        # Auto-fit column widths (simple heuristic)
        for i, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                except Exception:
                    val = ''
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"attendance_{start}_{end}.xlsx"
        resp = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=False, methods=['get'])
    def export_payout(self, request):
        """Export attendance matrix with payout row: last row contains salary = total_hours * hourly_rate

        Query params:
        - date_from (YYYY-MM-DD) required
        - date_to (YYYY-MM-DD) required
        - employees: optional comma-separated emp_id list; defaults to active employees
        - employee: optional single emp_id to export for a specific employee
        """
        if openpyxl is None:
            return Response({"error": "openpyxl is required to export Excel. Please pip install openpyxl."}, status=500)

        start = request.query_params.get('date_from')
        end = request.query_params.get('date_to')
        if not start or not end:
            return Response({"error": "Please provide date_from and date_to in YYYY-MM-DD"}, status=400)
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date()
            end_date = datetime.strptime(end, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        # Accept either `employee` (single emp_id) or `employees` (comma-separated)
        emp_single = request.query_params.get('employee')
        emp_param = request.query_params.get('employees')
        if emp_single and str(emp_single).strip().isdigit():
            employees = list(Employee.objects.filter(emp_id=int(emp_single)).order_by('emp_id'))
        elif emp_param:
            emp_ids = [int(x.strip()) for x in emp_param.split(',') if x.strip().isdigit()]
            employees = list(Employee.objects.filter(emp_id__in=emp_ids).order_by('emp_id'))
        else:
            employees = list(Employee.objects.filter(status='active').order_by('emp_id'))

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payout'

        # Header row: Date + for each employee a single column with "CheckIn - CheckOut"
        headers = ['Date']
        for e in employees:
            label = e.name or str(e.emp_id)
            headers.append(f"{label} (In - Out)")

        for col_idx, head in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=head)

        # Fill rows per date (reuse logic from export_excel)
        row = 2
        delta = (end_date - start_date).days
        current_time = datetime.now().time()
        today = timezone.now().date()

        # track totals per employee
        totals = {emp.emp_id: 0.0 for emp in employees}

        for d_offset in range(delta + 1):
            rdate = start_date + timedelta(days=d_offset)
            ws.cell(row=row, column=1, value=rdate.strftime('%Y-%m-%d'))

            for idx, emp in enumerate(employees):
                col = 2 + idx
                atts = Attendance.objects.filter(employee=emp).filter(
                    Q(date=rdate) | Q(check_in__date=rdate)
                )
                day_atts = atts.filter(check_in__date=rdate).order_by('check_in')
                if day_atts.exists():
                    atts = day_atts
                else:
                    atts = atts.order_by('check_in')

                check_in_val = "--:--"
                check_out_val = "--:--"
                if atts.exists():
                    first = atts.first()
                    last = atts.last()
                    if first.check_in:
                        try:
                            check_in_val = first.check_in.strftime('%I:%M %p')
                        except Exception:
                            check_in_val = str(first.check_in)
                    if last.check_out:
                        try:
                            check_out_val = last.check_out.strftime('%I:%M %p')
                        except Exception:
                            check_out_val = str(last.check_out)
                    # accumulate total hours for this day
                    totals[emp.emp_id] += sum(att.total_hours for att in atts)

                else:
                    # Leave handling
                    leave = PaidLeave.objects.filter(
                        employee=emp,
                        approved=True,
                        start_time__date__lte=rdate,
                        end_time__date__gte=rdate,
                    ).first()
                    if leave:
                        check_in_val = "Leave"
                        check_out_val = ""
                    else:
                        if not emp.start_time or not emp.end_time:
                            if rdate < today:
                                check_in_val = "Absent"
                                check_out_val = ""
                            else:
                                check_in_val = "--:--"
                                check_out_val = "--:--"
                        else:
                            if rdate < today or (rdate == today and current_time >= emp.end_time):
                                check_in_val = "Absent"
                                check_out_val = ""
                            else:
                                check_in_val = "--:--"
                                check_out_val = "--:--"

                ws.cell(row=row, column=col, value=f"{check_in_val} - {check_out_val}" if check_out_val else check_in_val)

            row += 1

        # Append totals row and salary row
        # blank separator
        row += 1
        totals_row = row
        ws.cell(row=totals_row, column=1, value='Total Hours')
        for idx, emp in enumerate(employees):
            col = 2 + idx
            hrs = round(totals.get(emp.emp_id, 0.0), 2)
            ws.cell(row=totals_row, column=col, value=hrs)

        # salary row
        row += 1
        salary_row = row
        ws.cell(row=salary_row, column=1, value='Salary')
        for idx, emp in enumerate(employees):
            col = 2 + idx
            hourly = float(emp.hourly_rate or 0)
            pay = round(totals.get(emp.emp_id, 0.0) * hourly, 2)
            ws.cell(row=salary_row, column=col, value=pay)

        # Auto-fit column widths
        for i, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                except Exception:
                    val = ''
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"payout_{start}_{end}.xlsx"
        resp = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=False, methods=['get'])
    def daily_report(self, request):
        """Get daily attendance report for a specific date (defaults to today)"""
        date_str = request.query_params.get('date')
        today = datetime.now().date()

        if not date_str:
            report_date = today
        else:
            try:
                report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({"error": "Invalid date format"}, status=400)

        attendances = Attendance.objects.filter(date=report_date)
        total_active_employees = Employee.objects.filter(status='active').count()

        # Count UNIQUE employees present today
        present_count = attendances.values('employee').distinct().count()
        total_hours = round(sum(att.total_hours for att in attendances), 2)

        active_employees = Employee.objects.filter(status='active')
        present_employee_ids = set(attendances.values_list('employee', flat=True).distinct())
        absent_details = []
        pending_count = 0
        current_time = datetime.now().time()

        for employee in active_employees:
            if employee.emp_id in present_employee_ids:
                continue

            if not employee.start_time or not employee.end_time:
                pending_count += 1
                continue

            if report_date < today or (report_date == today and current_time >= employee.end_time):
                # Check approved paid leaves covering this date
                on_leave = PaidLeave.objects.filter(
                    employee=employee,
                    approved=True,
                    start_time__date__lte=report_date,
                    end_time__date__gte=report_date,
                ).exists()

                status_val = "on_leave" if on_leave else "absent"
                absent_details.append({
                    "id": None,
                    "employee": employee.emp_id,
                    "employee_name": employee.name,
                    "date": report_date,
                    "check_in": None,
                    "check_out": None,
                    "message_late": None,
                    "status": status_val,
                    "total_hours": "0h 0m",
                    "is_late": False,
                    "created_at": None,
                    "updated_at": None,
                })
            else:
                pending_count += 1

        absent_count = len(absent_details)

        # Count how many UNIQUE employees were late at least once today
        late_count = attendances.filter(status='late').values('employee').distinct().count()
        on_time_count = present_count - late_count

        return Response({
            "date": report_date,
            "total_active_employees": total_active_employees,
            "present": present_count,
            "absent": absent_count,
            "pending": pending_count,
            "on_time": on_time_count,
            "late": late_count,
            "total_hours": format_hours_display(total_hours),
            "total_hours_value": total_hours,
            "attendance_details": AttendanceSerializer(attendances, many=True).data,
            "absent_details": absent_details
        })

    @action(detail=False, methods=['get'])
    def weekly_report(self, request):
        """Get weekly attendance report with rounded hours and unique counts"""
        today = datetime.now().date()
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)

        attendances = Attendance.objects.filter(date__range=[start_date, end_date])

        # FIX: Round the total hours to 2 decimal places
        total_hours = round(sum(att.total_hours for att in attendances), 2)
        
        # Count unique employee-day combinations
        total_working_records = attendances.values('employee', 'date').distinct().count()
        late_arrivals = attendances.filter(status='late').values('employee', 'date').distinct().count()

        return Response({
            "week": f"{start_date} to {end_date}",
            "total_records": total_working_records,
            "total_hours": format_hours_display(total_hours),
            "total_hours_value": total_hours,
            "late_arrivals": late_arrivals,
            "average_hours_per_day": round(total_hours / 7, 2) if total_working_records > 0 else 0
        })

    @action(detail=False, methods=['get'])
    def monthly_report(self, request):
        """Get monthly attendance report with rounded hours and unique counts"""
        today = datetime.now().date()
        month = int(request.query_params.get('month', today.month))
        year = int(request.query_params.get('year', today.year))

        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)

        attendances = Attendance.objects.filter(date__range=[start_date, end_date])

        # FIX: Round total hours
        total_hours = round(sum(att.total_hours for att in attendances), 2)
        
        # FIX: Count unique employee-day instances (Working Days)
        unique_working_days = attendances.values('employee', 'date').distinct().count()
        unique_employees = attendances.values('employee').distinct().count()
        late_arrivals = attendances.filter(status='late').values('employee', 'date').distinct().count()

        return Response({
            "month": f"{year}-{month:02d}",
            "total_working_days": unique_working_days,
            "total_hours_worked": format_hours_display(total_hours),
            "total_hours_worked_value": total_hours,
            "unique_employees": unique_employees,
            "late_arrivals": late_arrivals,
            "average_daily_attendance": round(unique_working_days / unique_employees, 2) if unique_employees > 0 else 0
        })

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def mark_absent(self, request):
        """Mark absent for employees who have no attendance and no approved leave.

        POST params:
        - date (YYYY-MM-DD) OR date_from/date_to
        - employees (optional comma-separated emp_id list) or employee (single emp_id)
        """
        # parse date range
        date_str = request.data.get('date') or request.query_params.get('date')
        start = request.data.get('date_from') or request.query_params.get('date_from')
        end = request.data.get('date_to') or request.query_params.get('date_to')

        if date_str:
            try:
                start_date = end_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({"error": "Invalid date format"}, status=400)
        elif start and end:
            try:
                start_date = datetime.strptime(start, '%Y-%m-%d').date()
                end_date = datetime.strptime(end, '%Y-%m-%d').date()
            except ValueError:
                return Response({"error": "Invalid date format"}, status=400)
        else:
            return Response({"error": "Provide date or date_from and date_to (YYYY-MM-DD)"}, status=400)

        # employees filter
        emp_single = request.data.get('employee') or request.query_params.get('employee')
        emp_param = request.data.get('employees') or request.query_params.get('employees')
        if emp_single and str(emp_single).strip().isdigit():
            employees = list(Employee.objects.filter(emp_id=int(emp_single)).order_by('emp_id'))
        elif emp_param:
            emp_ids = [int(x.strip()) for x in str(emp_param).split(',') if x.strip().isdigit()]
            employees = list(Employee.objects.filter(emp_id__in=emp_ids).order_by('emp_id'))
        else:
            employees = list(Employee.objects.filter(status='active').order_by('emp_id'))

        created = []
        skipped = []

        delta = (end_date - start_date).days
        for d_offset in range(delta + 1):
            rdate = start_date + timedelta(days=d_offset)
            for emp in employees:
                # If no shift times, still create absent record (use midnight as check_in)
                if not emp.start_time or not emp.end_time:
                    # but if an attendance exists or leave exists skip
                    exists = Attendance.objects.filter(employee=emp).filter(
                        Q(date=rdate) | Q(check_in__date=rdate)
                    ).exists()
                    on_leave = PaidLeave.objects.filter(
                        employee=emp,
                        approved=True,
                        start_time__date__lte=rdate,
                        end_time__date__gte=rdate,
                    ).exists()
                    if exists:
                        skipped.append({"employee": emp.emp_id, "date": rdate.isoformat(), "reason": "has_attendance"})
                        continue
                    if on_leave:
                        skipped.append({"employee": emp.emp_id, "date": rdate.isoformat(), "reason": "on_leave"})
                        continue
                    # create absent record with midnight check_in
                    check_dt = datetime.combine(rdate, time(0, 0))
                    att = Attendance.objects.create(
                        employee=emp,
                        date=rdate,
                        check_in=check_dt,
                        check_out=check_dt,
                        status='absent',
                        message_late='Marked absent by system'
                    )
                    created.append({"employee": emp.emp_id, "date": rdate.isoformat(), "id": att.id})
                    continue

                # if attendance exists for this date (either stored date or check_in date), skip
                exists = Attendance.objects.filter(employee=emp).filter(
                    Q(date=rdate) | Q(check_in__date=rdate)
                ).exists()
                if exists:
                    skipped.append({"employee": emp.emp_id, "date": rdate.isoformat(), "reason": "has_attendance"})
                    continue

                # if approved leave exists, skip
                on_leave = PaidLeave.objects.filter(
                    employee=emp,
                    approved=True,
                    start_time__date__lte=rdate,
                    end_time__date__gte=rdate,
                ).exists()
                if on_leave:
                    skipped.append({"employee": emp.emp_id, "date": rdate.isoformat(), "reason": "on_leave"})
                    continue

                # create absent attendance record with zero duration (check_in==check_out at shift start)
                check_dt = datetime.combine(rdate, emp.start_time)
                att = Attendance.objects.create(
                    employee=emp,
                    date=rdate,
                    check_in=check_dt,
                    check_out=check_dt,
                    status='absent',
                    message_late='Marked absent by system'
                )
                created.append({"employee": emp.emp_id, "date": rdate.isoformat(), "id": att.id})

        return Response({"created_count": len(created), "created": created, "skipped": skipped})

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def check_in(self, request):
        """Check in an employee"""
        emp_id = request.data.get('emp_id')
        
        if not emp_id:
            return Response(
                {"error": "emp_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            employee = Employee.objects.get(emp_id=emp_id)
        except Employee.DoesNotExist:
            return Response(
                {"error": f"Employee with id {emp_id} not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        today = timezone.now().date()
        now = timezone.now()

        # Consider the latest attendance record regardless of date so an open check-in
        # on a previous date will be treated as the next action's check-out.
        last_att = Attendance.objects.filter(employee=employee).order_by('-check_in').first()

        # If there's an open attendance (no check_out), treat this request as a check-out
        if last_att and last_att.check_out is None:
            duration = (now - last_att.check_in).total_seconds() / 3600
            if duration > 14:
                # Auto-check-out at shift end and create a new attendance (check-in)
                # Determine shift end datetime based on last_att.date and employee.end_time
                if getattr(employee, 'end_time', None):
                    shift_end_dt = datetime.combine(last_att.date, employee.end_time)
                    # If overnight shift, end_time may be on next day
                    if getattr(employee, 'start_time', None) and employee.end_time <= employee.start_time:
                        shift_end_dt += timedelta(days=1)
                else:
                    # Fallback: cap at 14 hours after check_in
                    shift_end_dt = last_att.check_in + timedelta(hours=14)

                # Ensure shift_end_dt is not in the future beyond 'now'
                if shift_end_dt > now:
                    shift_end_dt = now

                # Mark previous attendance as checked out at shift_end_dt and log the auto action in message_late
                prev_msg = last_att.message_late or ''
                try:
                    missing_date = last_att.check_in.date()
                except Exception:
                    missing_date = last_att.date
                missing_note = f"you haven't checked out {missing_date.strftime('%Y-%m-%d')}"
                last_att.check_out = shift_end_dt
                last_att.message_late = (prev_msg + ' | ' + missing_note).strip(' |')
                last_att.save()

                # Create a new attendance record representing the new check-in (current scan)
                new_status = 'on_time'
                new_late_msg = None
                if employee.start_time:
                    # Determine lateness for the new check-in against today's shift start
                    new_shift_start = datetime.combine(now.date(), employee.start_time)
                    if getattr(employee, 'end_time', None) and employee.end_time <= employee.start_time and new_shift_start > now:
                        new_shift_start -= timedelta(days=1)
                    is_late_new = now > new_shift_start
                    new_status = 'late' if is_late_new else 'on_time'
                    if is_late_new:
                        mins = int((now - new_shift_start).total_seconds() / 60)
                        new_late_msg = f"you are late {mins}m"

                new_att = Attendance.objects.create(
                    employee=employee,
                    date=now.date(),
                    check_in=now,
                    status=new_status,
                    message_late=new_late_msg
                )

                total_hours_prev = round(min((last_att.check_out - last_att.check_in).total_seconds() / 3600, 14.0), 2)

                return Response(
                    {
                        "message": "Auto check-out performed and new check-in created",
                        "previous_record": AttendanceSerializer(last_att).data,
                        "new_record": AttendanceSerializer(new_att).data,
                        "previous_total_hours": format_hours_display(total_hours_prev),
                        "previous_total_hours_value": total_hours_prev
                    },
                    status=status.HTTP_200_OK
                )

            # Normal check-out within 14 hours
            last_att.check_out = now
            last_att.save()

            total_hours = round(min((now - last_att.check_in).total_seconds() / 3600, 14.0), 2)

            return Response(
                {
                    "message": "Check-out successful",
                    "total_hours": format_hours_display(total_hours),
                    "total_hours_value": total_hours,
                    "record": AttendanceSerializer(last_att).data
                },
                status=status.HTTP_200_OK
            )

        # Otherwise create a new check-in. Determine lateness and present minutes-based message.
        status_val = 'on_time'
        late_msg = None
        if employee.start_time:
            shift_start_dt = datetime.combine(today, employee.start_time)
            if getattr(employee, 'end_time', None) and employee.end_time <= employee.start_time and shift_start_dt > now:
                shift_start_dt -= timedelta(days=1)

            is_late = now > shift_start_dt
            status_val = 'late' if is_late else 'on_time'
            if is_late:
                minutes_late = int((now - shift_start_dt).total_seconds() / 60)
                late_msg = f"you are late {minutes_late}m"

        attendance = Attendance.objects.create(
            employee=employee,
            date=today,
            check_in=now,
            status=status_val,
            message_late=late_msg
        )

        return Response(
            {
                "message": "Check-in successful",
                "record": AttendanceSerializer(attendance).data,
            },
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def auto_attendance(self, request):
        emp_id = request.data.get('emp_id')
        
        if not emp_id:
            return Response({"error": "emp_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(emp_id=emp_id)
        except Employee.DoesNotExist:
            return Response({"error": f"Employee with id {emp_id} not found"}, status=status.HTTP_404_NOT_FOUND)

        # UPDATED: Use system clock time (since USE_TZ = False)
        now = datetime.now() 
        today = now.date()
        current_time = now.time()
        
        # Look for the latest attendance record regardless of date so open check-ins carry over
        last_attendance = Attendance.objects.filter(employee=employee).order_by('-check_in').first()
        # For total hours calculation when checking out, use the last open check-in (the same record)
        first_checkin = None
        if last_attendance and last_attendance.check_out is None:
            first_checkin = last_attendance

        attendance_info = {
            "emp_id": employee.emp_id,
            "employee_name": employee.name,
            "profile_img": f"http://{settings.SERVER_IP}:{settings.SERVER_PORT}{employee.profile_img.url}" if employee.profile_img else None,
            "shift_type": employee.shift_type,
            "timestamp": now.strftime('%I:%M %p'), 
        }
        
        # Determine action: Check-in or Check-out
        # if not last_attendance or (last_attendance.check_out is not None):
        #     # ===== NEW CHECK-IN =====
        #     shift_start = employee.start_time
        #     is_late = current_time > shift_start
            
        #     status_val = 'late' if is_late else 'on_time'
            
        #     late_msg = "On time"
        #     if is_late:
        #         # UPDATED: No timezone awareness needed because USE_TZ = False
        #         shift_start_dt = datetime.combine(today, shift_start)
        #         minutes_late = int((now - shift_start_dt).total_seconds() / 60)
        #         late_msg = f"{minutes_late} minutes late"
            
        #     attendance = Attendance.objects.create(
        #         employee=employee,
        #         date=today,
        #         check_in=now, # Saves literal system time to DB
        #         message_late=late_msg,
        #         status=status_val
        #     )
            
        #     action = "check_in"
        #     message = "Check-in successful"
        #     attendance_info.update({
        #         "action": action,
        #         "check_in": now.strftime('%I:%M %p'),
        #         "check_out": "--:--",
        #         "is_late": is_late,
        #         "late_message": late_msg,
        #         "total_hours_today": 0
        #     })
            
        # Determine action: Check-in or Check-out
        if not last_attendance or (last_attendance.check_out is not None):
            # ===== NEW CHECK-IN =====
            shift_start = employee.start_time
            is_late = False
            late_msg = "On time"
            if shift_start:
                shift_start_dt = datetime.combine(today, shift_start)
                # Adjust for overnight shifts: if shift_end <= shift_start and shift_start is after now,
                # the actual shift start was yesterday.
                if getattr(employee, 'end_time', None) and employee.end_time <= shift_start and shift_start_dt > now:
                    shift_start_dt -= timedelta(days=1)

                is_late = now > shift_start_dt
                if is_late:
                    total_minutes = int((now - shift_start_dt).total_seconds() / 60)
                    # Logic to format as "1h 15m" or just "15m"
                    if total_minutes >= 60:
                        hours = total_minutes // 60
                        minutes = total_minutes % 60
                        if minutes > 0:
                            late_msg = f"{hours}h {minutes}m late"
                        else:
                            late_msg = f"{hours}h late"
                    else:
                        late_msg = f"{total_minutes}m late"
            
            status_val = 'late' if is_late else 'on_time'

            attendance = Attendance.objects.create(
                employee=employee,
                date=today,
                check_in=now,
                message_late=late_msg,
                status=status_val
            )
            
            action = "check_in"
            message = "Check-in successful"
            attendance_info.update({
                "action": action,
                "check_in": now.strftime('%I:%M %p'),
                "check_out": "--:--",
                "is_late": is_late,
                "late_message": late_msg,
                "total_hours_today": "0h 0m",
                "total_hours_today_value": 0
            })
        elif last_attendance.check_in is not None and last_attendance.check_out is None:
            # ===== CHECK-OUT =====
            # Both are now "naive" datetimes, so math works perfectly
            duration = (now - last_attendance.check_in).total_seconds() / 3600
            if duration > 14:
                # Auto-check-out at shift end and create a new check-in record
                if getattr(employee, 'end_time', None):
                    shift_end_dt = datetime.combine(last_attendance.date, employee.end_time)
                    if getattr(employee, 'start_time', None) and employee.end_time <= employee.start_time:
                        shift_end_dt += timedelta(days=1)
                else:
                    shift_end_dt = last_attendance.check_in + timedelta(hours=14)

                if shift_end_dt > now:
                    shift_end_dt = now

                prev_msg = last_attendance.message_late or ''
                try:
                    missing_date = last_attendance.check_in.date()
                except Exception:
                    missing_date = last_attendance.date
                missing_note = f"you haven't checked out {missing_date.strftime('%Y-%m-%d')}"
                last_attendance.check_out = shift_end_dt
                last_attendance.message_late = (prev_msg + ' | ' + missing_note).strip(' |')
                last_attendance.save()

                # create new attendance as new check-in
                new_status = 'on_time'
                new_late_msg = None
                if employee.start_time:
                    new_shift_start = datetime.combine(now.date(), employee.start_time)
                    if getattr(employee, 'end_time', None) and employee.end_time <= employee.start_time and new_shift_start > now:
                        new_shift_start -= timedelta(days=1)
                    is_late_new = now > new_shift_start
                    new_status = 'late' if is_late_new else 'on_time'
                    if is_late_new:
                        mins = int((now - new_shift_start).total_seconds() / 60)
                        new_late_msg = f"you are late {mins}m"

                new_att = Attendance.objects.create(
                    employee=employee,
                    date=now.date(),
                    check_in=now,
                    status=new_status,
                    message_late=new_late_msg
                )

                total_hours = round(min((last_attendance.check_out - last_attendance.check_in).total_seconds() / 3600, 14.0), 2)

                action = "auto_check_out_and_check_in"
                message = "Auto check-out performed and new check-in created"
                attendance_info.update({
                    "previous_record": AttendanceSerializer(last_attendance).data,
                    "new_record": AttendanceSerializer(new_att).data,
                    "previous_total_hours": format_hours_display(total_hours),
                    "previous_total_hours_value": total_hours
                })
            else:
                last_attendance.check_out = now # Saves literal system time to DB
                last_attendance.save()
                
                total_hours = 0
                if first_checkin:
                    total_duration = (now - first_checkin.check_in).total_seconds() / 3600
                    total_hours = round(min(total_duration, 14.0), 2)
                
                action = "check_out"
                message = "Check-out successful"
                attendance_info.update({
                    "action": action,
                    # UPDATED: Just use the time in DB directly (it's already local)
                    "check_in": last_attendance.check_in.strftime('%I:%M %p'),
                    "check_out": now.strftime('%I:%M %p'),
                    "is_late": False,
                    "late_message": None,
                    "total_hours_today": format_hours_display(total_hours),
                    "total_hours_today_value": total_hours
                })

        response_payload = {
            "message": message,
            "action": action,
            "data": attendance_info
        }

        # Broadcast to WebSocket
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "biometric_device",
                {
                    "type": "biometric_event",
                    "data": attendance_info 
                }
            )
        except Exception as e:
            print(f"WS Error: {e}")

        return Response(response_payload, status=status.HTTP_200_OK)

class PaidLeaveViewSet(viewsets.ModelViewSet):
    queryset = PaidLeave.objects.all().order_by('-start_time')
    serializer_class = PaidLeaveSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a leave request"""
        leave = self.get_object()
        leave.approved = True
        leave.approved_by = request.data.get('approved_by', 'Admin')
        leave.save()
        return Response(
            {"message": "Leave approved", "leave": PaidLeaveSerializer(leave).data}
        )

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a leave request"""
        leave = self.get_object()
        leave.delete()
        return Response(
            {"message": "Leave request rejected"},
            status=status.HTTP_204_NO_CONTENT
        )

    @action(detail=False, methods=['get'])
    def pending_approvals(self, request):
        """Get all pending leave requests"""
        pending = PaidLeave.objects.filter(approved=False)
        serializer = self.get_serializer(pending, many=True)
        return Response({
            "pending_count": pending.count(),
            "leaves": serializer.data
        })

    @action(detail=False, methods=['get'])
    def employee_leaves(self, request):
        """Get leaves for a specific employee"""
        emp_id = request.query_params.get('emp_id')
        if not emp_id:
            return Response(
                {"error": "emp_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        leaves = PaidLeave.objects.filter(employee__emp_id=emp_id)
        serializer = self.get_serializer(leaves, many=True)
        return Response(serializer.data)

class ShiftViewSet(viewsets.ModelViewSet):
    queryset = Shift.objects.all()
    serializer_class = ShiftSerializer
    permission_classes = [IsAuthenticated]